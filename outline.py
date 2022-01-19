import cv2
import numpy as np
import glob
import csv

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, LineChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font, RichTextProperties
from openpyxl.drawing.line import LineProperties
from copy import deepcopy

def drawing(file_path, save_path):
    img = cv2.imread(file_path)
    flip = cv2.flip(img, 0)
    gray = cv2.cvtColor(flip, cv2.COLOR_BGR2GRAY)
    ret, bin_img = cv2.threshold(gray, 20, 255, cv2.THRESH_BINARY)

    # 点数が多い
    contours, hierarchy = cv2.findContours(bin_img, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_NONE)
    # 点数が少ない
    # contours, hierarchy = cv2.findContours(bin_img, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_TC89_L1)
    contours = list(filter(lambda x: cv2.contourArea(x) > 100, contours))

    cv2.drawContours(img, contours, -1, color=(0, 0, 255), thickness=2)

    # cv2.imshow("image", img)
    # cv2.waitKey(0)

    contours = [np.squeeze(cnt, axis=1) for cnt in contours]
    for i, cnt in enumerate(contours):
        np.savetxt(save_path + '\\contour_{}.csv'.format(i), cnt, delimiter=',')

def csv_treat(save_path):
    wb = openpyxl.Workbook()
    ws_data = wb.create_sheet(title='data_sheet')
    ws_chart = wb.create_sheet(title='chart_sheet')
    wb.remove(wb.worksheets[0])
    file = glob.glob(save_path + '\\*.csv')

    chart = ScatterChart()
    num_shape = 0
    for i in file:
        csv_file = open(i, "r", encoding="ms932", errors="", newline="" )
        f = csv.reader(csv_file, delimiter=",", doublequote=True, lineterminator="\r\n", quotechar='"', skipinitialspace=True)
        ws_data.cell(1, num_shape*2+1, value = "object#{}".format(num_shape))


        num_row = 0
        for row in f:
            row1_flt = float(row[0].split("E")[0])
            row2_flt = float(row[1].split("E")[0])
            ws_data.cell(num_row+2, num_shape*2+1, value = row1_flt)
            ws_data.cell(num_row+2, num_shape*2+2, value = row2_flt)

            num_row += 1

        xvalues = Reference(ws_data, min_col=num_shape*2+1, min_row=2, max_row=num_row)
        values = Reference(ws_data, min_col=num_shape*2+2, min_row=2, max_row=num_row)
        series = Series(values, xvalues, title="object#{}".format(num_shape))
        chart.series.append(series)

        print(num_shape)
        num_shape += 1

    ws_chart.add_chart(chart, "A1")

    wb.save('confirm.xlsx')

def main():
    file_path = "C:\\Users\FUTOSHI\Dropbox\Futoshi_Mac\Python_PJ\\1_OUTLINE\\test_1.jpg"
    save_path = "C:\\Users\FUTOSHI\Dropbox\Futoshi_Mac\Python_PJ\\1_OUTLINE\csv"
    drawing(file_path, save_path)
    csv_treat(save_path)

if __name__ == "__main__":
    main()